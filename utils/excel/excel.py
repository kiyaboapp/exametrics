import asyncio
import os
import shutil
from datetime import datetime
import time
import logging
from typing import Any, Dict
import aiomysql
from dotenv import load_dotenv
import numpy as np
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.sql import text
from uuid6 import uuid6

# Load environment variables
if not os.path.exists('.env'):
    raise FileNotFoundError("Missing .env file in project root")
load_dotenv()

# Configure logging
logging.basicConfig(
    level=getattr(logging, os.getenv('LOG_LEVEL', 'INFO')),
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Database configuration with connection pool
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': int(os.getenv('DB_PORT', 3306)),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', ''),
    'db': os.getenv('DB_NAME', 'exametrics'),
    'maxsize': int(os.getenv('DB_POOL_SIZE', 5))
}



async def get_excel_workbook_name(
    ward_name: str = "",
    council_name: str = "",
    region_name: str = "",
    school_type: str = "",
    practical_mode: int = 0
) -> str:
    """Generate a unique Excel workbook filename based on parameters and timestamp."""
    separator = "_"
    result = []
    
    # Build base name from non-empty parameters
    if ward_name:
        result.append(ward_name.replace(" ", "_"))
    if council_name:
        result.append(council_name.replace(" ", "_"))
    if region_name:
        result.append(region_name.replace(" ", "_"))
    if school_type:
        result.append(school_type.replace(" ", "_"))
    
    # Handle practical mode
    if practical_mode != 0:
        mode_name = {
            2: "Practical",
            1: "Theory",
            0: "Theory_Practical"
        }.get(practical_mode, "CustomMode")
        result.append(mode_name)
    
    # Add timestamp for uniqueness
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Final assembly
    if result:
        return f"ENTRY_{separator.join(result)}_{timestamp}".upper()
    return f"ENTRY_SHEET_JOINT_EXAM_{timestamp}".upper()

async def export_to_excel_Old(
    exam_id: str,
    ward_name: str = "",
    council_name: str = "",
    region_name: str = "",
    school_type: str = "",
    practical_mode: int = 0,
    marks_filler:str=""
) -> str:
    """Export student data to an Excel file and return the file path."""
    pool = None
    conn = None
    cursor = None
    
    try:
        pool = await aiomysql.create_pool(**DB_CONFIG)
        conn = await pool.acquire()
        cursor = await conn.cursor(aiomysql.DictCursor)
        
        # Debug: Check data existence
        await cursor.execute(
            "SELECT COUNT(*) AS count FROM exam_subjects WHERE exam_id = %s",
            (exam_id,)
        )
        exam_subjects_count = (await cursor.fetchone())['count']
        logger.info(f"Found {exam_subjects_count} subjects for exam_id={exam_id}")
        
        await cursor.execute(
            "SELECT COUNT(*) AS count FROM student_subjects WHERE exam_id = %s",
            (exam_id,)
        )
        student_subjects_count = (await cursor.fetchone())['count']
        logger.info(f"Found {student_subjects_count} student-subject mappings for exam_id={exam_id}")
        
        # Build WHERE clause with parameterized query
        where_clause = "WHERE es.exam_id = %s"
        params = [exam_id]
        location_filter = []
        if ward_name:
            location_filter.append("s.ward_name = %s")
            params.append(ward_name)
        if council_name:
            location_filter.append("s.council_name = %s")
            params.append(council_name)
        if region_name:
            location_filter.append("s.region_name = %s")
            params.append(region_name)
        if school_type:
            location_filter.append("s.school_type = %s")
            params.append(school_type)
        
        if location_filter:
            where_clause += " AND " + " AND ".join(location_filter)
        
        if practical_mode == 2:  # ExportOnlyPracticalVersions
            where_clause += " AND es.has_practical = TRUE"
        
        # SQL query with INNER JOINs
        sql = f"""
        SELECT st.student_id, st.student_global_id, st.full_name, st.sex,
               es.subject_code, es.subject_name, es.subject_short,
               s.centre_number, s.school_name,
               s.ward_name, s.council_name, s.region_name,
               s.school_type, es.has_practical
        FROM schools s
        INNER JOIN students st ON s.centre_number = st.centre_number
        INNER JOIN student_subjects ss ON st.student_global_id = ss.student_global_id AND st.exam_id = ss.exam_id
        INNER JOIN exam_subjects es ON ss.exam_id = es.exam_id AND ss.subject_code = es.subject_code
        {where_clause}
        ORDER BY st.student_id ASC, es.subject_code ASC
        """
        
        await cursor.execute(sql, params)
        records = await cursor.fetchall()
        
        # Error message for empty result
        msg = "NO STUDENT REGISTERED FOR THIS CONDITION!"
        if ward_name:
            msg += f" WARD_NAME={ward_name}"
        if council_name:
            msg += f" COUNCIL={council_name}"
        if region_name:
            msg += f" REGION={region_name}"
        if school_type:
            msg += f" SCHOOL TYPE={school_type}"
        if practical_mode != 0:
            msg += f" PRACTICAL MODE={practical_mode}"
        if exam_id:
            msg += f" EXAM_ID={exam_id}"
        
        if not records:
            logger.error(f"Error 1002: {msg}")
            raise ValueError(msg)
        
        # Initialize dictionaries for subjects and schools
        subject_dict = {}
        school_dict = {}
        data_array = []
        
        # Process records
        prev=None
        for record in records:
            prev != record['centre_number'] and logger.info(f"centre_number changed: {record['centre_number']}") or True; prev = record['centre_number']
            school_key = (
                f"{record['centre_number'] or ''}|{record['school_name'] or ''}|"
                f"{record['ward_name'] or ''}|{record['council_name'] or ''}|"
                f"{record['region_name'] or ''}|{record['school_type'] or ''}"
            )
            # logger.info(school_key)
            
            if practical_mode != 2:  # Not ExportOnlyPracticalVersions
                # Add theory row
                data_array.append([
                    record['student_id'] or '',  # Use student_id for readability
                    record['full_name'] or '',
                    record['sex'] or '',
                    '',  # Empty marks column
                    str(record['subject_code'] or ''),
                    record['subject_name'] or '',
                    record['subject_short'] or '',
                    record['centre_number'] or '',
                    record['school_name'] or '',
                    record['ward_name'] or '',
                    record['council_name'] or '',
                    record['region_name'] or '',
                    exam_id,
                    record['student_global_id'] or ''  # Add student_global_id in column N
                ])
                
                subj_key = (
                    f"{record['subject_code'] or ''}|{record['subject_name'] or ''}|"
                    f"{record['subject_short'] or ''}"
                )
                if subj_key not in subject_dict:
                    subject_dict[subj_key] = [
                        str(record['subject_code'] or ''),
                        record['subject_name'] or '',
                        record['subject_short'] or '',
                        0, 0
                    ]
                subject_dict[subj_key][3] += 1
                
                if school_key not in school_dict:
                    school_dict[school_key] = [
                        record['centre_number'] or '',
                        record['school_name'] or '',
                        record['ward_name'] or '',
                        record['council_name'] or '',
                        record['region_name'] or '',
                        0, 0
                    ]
                school_dict[school_key][5] += 1
            
            if record['has_practical'] and practical_mode != 1:  # Not ExportWithoutPractical
                # Add practical row
                data_array.append([
                    record['student_id'] or '',  # Use student_id for readability
                    record['full_name'] or '',
                    record['sex'] or '',
                    '',  # Empty marks column
                    f"{record['subject_code'] or ''}-P",
                    f"{record['subject_name'] or ''}-Practical",
                    f"{record['subject_short'] or ''}-P",
                    record['centre_number'] or '',
                    record['school_name'] or '',
                    record['ward_name'] or '',
                    record['council_name'] or '',
                    record['region_name'] or '',
                    exam_id,
                    record['student_global_id'] or ''  # Add student_global_id in column N
                ])
                
                prac_subj_key = (
                    f"{record['subject_code'] or ''}-P|"
                    f"{record['subject_name'] or ''}-Practical|"
                    f"{record['subject_short'] or ''}-P"
                )
                if prac_subj_key not in subject_dict:
                    subject_dict[prac_subj_key] = [
                        f"{record['subject_code'] or ''}-P",
                        f"{record['subject_name'] or ''}-Practical",
                        f"{record['subject_short'] or ''}-P",
                        0, 0
                    ]
                subject_dict[prac_subj_key][3] += 1
                
                if school_key not in school_dict:
                    school_dict[school_key] = [
                        record['centre_number'] or '',
                        record['school_name'] or '',
                        record['ward_name'] or '',
                        record['council_name'] or '',
                        record['region_name'] or '',
                        0, 0
                    ]
                school_dict[school_key][5] += 1
        
        # Copy master.xlsm to output
        original_file = os.path.join(os.path.dirname(__file__), "master.xlsm")
        if not os.path.exists(original_file):
            logger.error("Error 1001: master.xlsm not found!")
            raise FileNotFoundError("master.xlsm not found!")
        
        os.makedirs("./output", exist_ok=True)
        save_path = os.path.join(
            "./output",
            f"{await get_excel_workbook_name(ward_name, council_name, region_name, school_type, practical_mode)}.xlsm"
        )
        
        shutil.copyfile(original_file, save_path)
        
        # Open Excel workbook
        workbook = openpyxl.load_workbook(save_path, keep_vba=True)
        students_sheet = workbook["Students"]
        subjects_sheet = workbook["Subjects"]
        schools_sheet = workbook["Schools"]
        interface_sheet = workbook["Interface"]
        
        # Clear Interface!A500
        if marks_filler=="":
            interface_sheet["A500"] = ""
        else:
            interface_sheet["A500"]=marks_filler
        
        # Set number formats
        students_sheet.column_dimensions["E"].number_format = "@"
        subjects_sheet.column_dimensions["A"].number_format = "@"
        
        # Write student data
        if data_array:
            for i, row in enumerate(data_array, start=2):
                for j, value in enumerate(row, start=1):
                    students_sheet.cell(row=i, column=j).value = value
            table1 = students_sheet.tables.get("Table1")
            if table1:
                table1.ref = f"A1:N{len(data_array) + 1}"  # Updated to include column N
        else:
            students_sheet["A2:N2"] = [""] * 14  # Updated to include column N
            table1 = students_sheet.tables.get("Table1")
            if table1:
                table1.ref = "A1:N2"
        
        # Write subjects data
        row_index = 2
        for subj_data in subject_dict.values():
            subjects_sheet[f"A{row_index}"] = subj_data[0]
            subjects_sheet[f"B{row_index}"] = subj_data[2]
            subjects_sheet[f"C{row_index}"] = subj_data[1]
            subjects_sheet[f"D{row_index}"] = f"=COUNTIFS(Students!E:E,A{row_index})"
            subjects_sheet[f"E{row_index}"] = f"=COUNTIFS(Students!E:E,A{row_index},Students!D:D,\">=0\")"
            subjects_sheet[f"F{row_index}"] = f"=D{row_index}-E{row_index}"
            subjects_sheet[f"G{row_index}"] = f"=IF(D{row_index}=0,\"N/A\",E{row_index}/D{row_index})"
            row_index += 1
        
        if row_index > 2:
            subjects_sheet[f"C{row_index}"] = "Total"
            subjects_sheet[f"D{row_index}"] = f"=SUM(D2:D{row_index-1})"
            subjects_sheet[f"E{row_index}"] = f"=SUM(E2:E{row_index-1})"
            subjects_sheet[f"F{row_index}"] = f"=SUM(F2:F{row_index-1})"
            subjects_sheet[f"G{row_index}"] = f"=IF(D{row_index}=0,\"N/A\",E{row_index}/D{row_index})"
        
        subjects_sheet.column_dimensions["G"].number_format = "0.0%"
        
        # Write schools data
        row_index = 2
        for school_data in school_dict.values():
            schools_sheet[f"A{row_index}"] = school_data[0]
            schools_sheet[f"B{row_index}"] = school_data[1]
            schools_sheet[f"C{row_index}"] = school_data[2]  # ward_name
            schools_sheet[f"D{row_index}"] = school_data[3]  # council_name
            schools_sheet[f"E{row_index}"] = school_data[4]  # region_name
            schools_sheet[f"F{row_index}"] = f"=COUNTIFS(Students!H:H,A{row_index})"
            schools_sheet[f"G{row_index}"] = f"=COUNTIFS(Students!H:H,A{row_index},Students!D:D,\">=0\")"
            schools_sheet[f"H{row_index}"] = f"=F{row_index}-G{row_index}"
            schools_sheet[f"I{row_index}"] = f"=IF(F{row_index}=0,\"N/A\",G{row_index}/F{row_index})"
            row_index += 1
        
        if row_index > 2:
            schools_sheet[f"E{row_index}"] = "Total"
            schools_sheet[f"F{row_index}"] = f"=SUM(F2:F{row_index-1})"
            schools_sheet[f"G{row_index}"] = f"=SUM(G2:G{row_index-1})"
            schools_sheet[f"H{row_index}"] = f"=SUM(H2:H{row_index-1})"
            schools_sheet[f"I{row_index}"] = f"=IF(F{row_index}=0,\"N/A\",G{row_index}/F{row_index})"
        
        schools_sheet.column_dimensions["I"].number_format = "0.0%"
        
        # Create dropdowns in Interface sheet
        centre_numbers = [str(school_data[0]) for school_data in school_dict.values() if school_data[0]]
        if centre_numbers:
            dv_centre = DataValidation(type="list", formula1=f'"{",".join(centre_numbers)}"', allow_blank=True)
            dv_centre.add("C5")
            interface_sheet.add_data_validation(dv_centre)
            logger.debug(f"Added dropdown to Interface!C5 with {len(centre_numbers)} centre numbers")
        
        subject_codes = [str(subj_data[0]) for subj_data in subject_dict.values() if subj_data[0]]
        if subject_codes:
            interface_sheet["C6"].number_format = "@"
            dv_subject = DataValidation(type="list", formula1=f'"{",".join(subject_codes)}"', allow_blank=True)
            dv_subject.add("C6")
            interface_sheet.add_data_validation(dv_subject)
            logger.debug(f"Added dropdown to Interface!C6 with {len(subject_codes)} subject codes")
        
        # Save and close
        workbook.save(save_path)
        workbook.close()
        
        logger.info(f"Export completed successfully to {save_path}")
        return save_path
    
    except aiomysql.Error as e:
        logger.error(f"Error 1004: Database error: {e}")
        raise
    except Exception as e:
        logger.error(f"Error 1003: Excel operation failed: {e}")
        raise
    finally:
        if cursor:
            await cursor.close()
        if conn and pool:
            pool.release(conn)
        if pool:
            pool.close()
            await pool.wait_closed()


async def export_to_excel(
    exam_id: str,
    ward_name: str = "",
    council_name: str = "",
    region_name: str = "",
    school_type: str = "",
    practical_mode: int = 0,
    marks_filler: str = ""
) -> str:
    """Export student data to an Excel file and return the file path."""
    pool = None
    conn = None
    cursor = None
    
    try:
        pool = await aiomysql.create_pool(**DB_CONFIG)
        conn = await pool.acquire()
        cursor = await conn.cursor(aiomysql.DictCursor)
        
        # Debug: Check data existence
        await cursor.execute(
            "SELECT COUNT(*) AS count FROM exam_subjects WHERE exam_id = %s",
            (exam_id,)
        )
        exam_subjects_count = (await cursor.fetchone())['count']
        logger.info(f"Found {exam_subjects_count} subjects for exam_id={exam_id}")
        
        await cursor.execute(
            "SELECT COUNT(*) AS count FROM student_subjects WHERE exam_id = %s",
            (exam_id,)
        )
        student_subjects_count = (await cursor.fetchone())['count']
        logger.info(f"Found {student_subjects_count} student-subject mappings for exam_id={exam_id}")
        
        # Build WHERE clause with parameterized query
        where_clause = "WHERE es.exam_id = %s"
        params = [exam_id]
        location_filter = []
        if ward_name:
            location_filter.append("s.ward_name = %s")
            params.append(ward_name)
        if council_name:
            location_filter.append("s.council_name = %s")
            params.append(council_name)
        if region_name:
            location_filter.append("s.region_name = %s")
            params.append(region_name)
        if school_type:
            location_filter.append("s.school_type = %s")
            params.append(school_type)
        
        if location_filter:
            where_clause += " AND " + " AND ".join(location_filter)
        
        if practical_mode == 2:  # ExportOnlyPracticalVersions
            where_clause += " AND es.has_practical = TRUE"
        
        # SQL query with INNER JOINs
        sql = f"""
        SELECT st.student_id, st.student_global_id, st.full_name, st.sex,
               es.subject_code, es.subject_name, es.subject_short,
               s.centre_number, s.school_name,
               s.ward_name, s.council_name, s.region_name,
               s.school_type, es.has_practical
        FROM schools s
        INNER JOIN students st ON s.centre_number = st.centre_number
        INNER JOIN student_subjects ss ON st.student_global_id = ss.student_global_id AND st.exam_id = ss.exam_id
        INNER JOIN exam_subjects es ON ss.exam_id = es.exam_id AND ss.subject_code = es.subject_code
        {where_clause}
        ORDER BY st.student_id ASC, es.subject_code ASC
        """
        
        await cursor.execute(sql, params)
        records = await cursor.fetchall()
        
        # Error message for empty result
        msg = "NO STUDENT REGISTERED FOR THIS CONDITION!"
        if ward_name:
            msg += f" WARD_NAME={ward_name}"
        if council_name:
            msg += f" COUNCIL={council_name}"
        if region_name:
            msg += f" REGION={region_name}"
        if school_type:
            msg += f" SCHOOL TYPE={school_type}"
        if practical_mode != 0:
            msg += f" PRACTICAL MODE={practical_mode}"
        if exam_id:
            msg += f" EXAM_ID={exam_id}"
        
        if not records:
            logger.error(f"Error 1002: {msg}")
            raise ValueError(msg)
        
        # Initialize dictionaries for subjects and schools
        subject_dict = {}
        school_dict = {}
        data_array = []
        
        # Process records
        prev = None
        for record in records:
            prev != record['centre_number'] and logger.info(f"centre_number changed: {record['centre_number']}") or True
            prev = record['centre_number']
            
            school_key = (
                f"{record['centre_number'] or ''}|{record['school_name'] or ''}|"
                f"{record['ward_name'] or ''}|{record['council_name'] or ''}|"
                f"{record['region_name'] or ''}|{record['school_type'] or ''}"
            )
            
            if practical_mode != 2:  # Not ExportOnlyPracticalVersions
                # Add theory row
                data_array.append([
                    record['student_id'] or '',
                    record['full_name'] or '',
                    record['sex'] or '',
                    '',
                    str(record['subject_code'] or ''),
                    record['subject_name'] or '',
                    record['subject_short'] or '',
                    record['centre_number'] or '',
                    record['school_name'] or '',
                    record['ward_name'] or '',
                    record['council_name'] or '',
                    record['region_name'] or '',
                    exam_id,
                    record['student_global_id'] or ''
                ])
                
                subj_key = (
                    f"{record['subject_code'] or ''}|{record['subject_name'] or ''}|"
                    f"{record['subject_short'] or ''}"
                )
                if subj_key not in subject_dict:
                    subject_dict[subj_key] = [
                        str(record['subject_code'] or ''),
                        record['subject_name'] or '',
                        record['subject_short'] or '',
                        0, 0
                    ]
                subject_dict[subj_key][3] += 1
                
                if school_key not in school_dict:
                    school_dict[school_key] = [
                        record['centre_number'] or '',
                        record['school_name'] or '',
                        record['ward_name'] or '',
                        record['council_name'] or '',
                        record['region_name'] or '',
                        0, 0
                    ]
                school_dict[school_key][5] += 1
            
            if record['has_practical'] and practical_mode != 1:  # Not ExportWithoutPractical
                # Add practical row
                data_array.append([
                    record['student_id'] or '',
                    record['full_name'] or '',
                    record['sex'] or '',
                    '',
                    f"{record['subject_code'] or ''}-P",
                    f"{record['subject_name'] or ''}-Practical",
                    f"{record['subject_short'] or ''}-P",
                    record['centre_number'] or '',
                    record['school_name'] or '',
                    record['ward_name'] or '',
                    record['council_name'] or '',
                    record['region_name'] or '',
                    exam_id,
                    record['student_global_id'] or ''
                ])
                
                prac_subj_key = (
                    f"{record['subject_code'] or ''}-P|"
                    f"{record['subject_name'] or ''}-Practical|"
                    f"{record['subject_short'] or ''}-P"
                )
                if prac_subj_key not in subject_dict:
                    subject_dict[prac_subj_key] = [
                        f"{record['subject_code'] or ''}-P",
                        f"{record['subject_name'] or ''}-Practical",
                        f"{record['subject_short'] or ''}-P",
                        0, 0
                    ]
                subject_dict[prac_subj_key][3] += 1
                
                if school_key not in school_dict:
                    school_dict[school_key] = [
                        record['centre_number'] or '',
                        record['school_name'] or '',
                        record['ward_name'] or '',
                        record['council_name'] or '',
                        record['region_name'] or '',
                        0, 0
                    ]
                school_dict[school_key][5] += 1
        
        # Copy master.xlsm to output
        original_file = os.path.join(os.path.dirname(__file__), "master.xlsm")
        if not os.path.exists(original_file):
            logger.error("Error 1001: master.xlsm not found!")
            raise FileNotFoundError("master.xlsm not found!")
        
        os.makedirs("./output", exist_ok=True)
        save_path = os.path.join(
            "./output",
            f"{await get_excel_workbook_name(ward_name, council_name, region_name, school_type, practical_mode)}.xlsm"
        )
        
        shutil.copyfile(original_file, save_path)
        
        # Open Excel workbook
        workbook = openpyxl.load_workbook(save_path, keep_vba=True)
        students_sheet = workbook["Students"]
        subjects_sheet = workbook["Subjects"]
        schools_sheet = workbook["Schools"]
        interface_sheet = workbook["Interface"]
        
        # Clear Interface!A500 or set marks_filler
        interface_sheet["A500"] = marks_filler if marks_filler else ""
        
        # Set number formats
        students_sheet.column_dimensions["E"].number_format = "@"
        subjects_sheet.column_dimensions["A"].number_format = "@"
        
        # Write student data using chunks to avoid memory issues
        chunk_size = 1000
        if data_array:
            for chunk_start in range(0, len(data_array), chunk_size):
                chunk = data_array[chunk_start:chunk_start + chunk_size]
                for i, row in enumerate(chunk, start=2 + chunk_start):
                    for j, value in enumerate(row, start=1):
                        students_sheet.cell(row=i, column=j).value = value
            table1 = students_sheet.tables.get("Table1")
            if table1:
                table1.ref = f"A1:N{len(data_array) + 1}"  # Updated to include column N
        else:
            students_sheet["A2:N2"] = [""] * 14  # Updated to include column N
            table1 = students_sheet.tables.get("Table1")
            if table1:
                table1.ref = "A1:N2"
        
        # Write subjects data
        row_index = 2
        for subj_data in subject_dict.values():
            subjects_sheet[f"A{row_index}"] = subj_data[0]
            subjects_sheet[f"B{row_index}"] = subj_data[2]
            subjects_sheet[f"C{row_index}"] = subj_data[1]
            subjects_sheet[f"D{row_index}"] = f"=COUNTIFS(Students!E:E,A{row_index})"
            subjects_sheet[f"E{row_index}"] = f"=COUNTIFS(Students!E:E,A{row_index},Students!D:D,\">=0\")"
            subjects_sheet[f"F{row_index}"] = f"=D{row_index}-E{row_index}"
            subjects_sheet[f"G{row_index}"] = f"=IF(D{row_index}=0,\"N/A\",E{row_index}/D{row_index})"
            row_index += 1
        
        if row_index > 2:
            subjects_sheet[f"C{row_index}"] = "Total"
            subjects_sheet[f"D{row_index}"] = f"=SUM(D2:D{row_index-1})"
            subjects_sheet[f"E{row_index}"] = f"=SUM(E2:E{row_index-1})"
            subjects_sheet[f"F{row_index}"] = f"=SUM(F2:F{row_index-1})"
            subjects_sheet[f"G{row_index}"] = f"=IF(D{row_index}=0,\"N/A\",E{row_index}/D{row_index})"
        
        subjects_sheet.column_dimensions["G"].number_format = "0.0%"
        
        # Write schools data
        row_index = 2
        for school_data in school_dict.values():
            schools_sheet[f"A{row_index}"] = school_data[0]
            schools_sheet[f"B{row_index}"] = school_data[1]
            schools_sheet[f"C{row_index}"] = school_data[2]  # ward_name
            schools_sheet[f"D{row_index}"] = school_data[3]  # council_name
            schools_sheet[f"E{row_index}"] = school_data[4]  # region_name
            schools_sheet[f"F{row_index}"] = f"=COUNTIFS(Students!H:H,A{row_index})"
            schools_sheet[f"G{row_index}"] = f"=COUNTIFS(Students!H:H,A{row_index},Students!D:D,\">=0\")"
            schools_sheet[f"H{row_index}"] = f"=F{row_index}-G{row_index}"
            schools_sheet[f"I{row_index}"] = f"=IF(F{row_index}=0,\"N/A\",G{row_index}/F{row_index})"
            row_index += 1
        
        if row_index > 2:
            schools_sheet[f"E{row_index}"] = "Total"
            schools_sheet[f"F{row_index}"] = f"=SUM(F2:F{row_index-1})"
            schools_sheet[f"G{row_index}"] = f"=SUM(G2:G{row_index-1})"
            schools_sheet[f"H{row_index}"] = f"=SUM(H2:H{row_index-1})"
            schools_sheet[f"I{row_index}"] = f"=IF(F{row_index}=0,\"N/A\",G{row_index}/F{row_index})"
        
        schools_sheet.column_dimensions["I"].number_format = "0.0%"
        
        # Create dropdowns in Interface sheet
        centre_numbers = [str(school_data[0]) for school_data in school_dict.values() if school_data[0]]
        if centre_numbers:
            dv_centre = DataValidation(type="list", formula1=f'"{",".join(centre_numbers)}"', allow_blank=True)
            dv_centre.add("C5")
            interface_sheet.add_data_validation(dv_centre)
            logger.debug(f"Added dropdown to Interface!C5 with {len(centre_numbers)} centre numbers")
        
        subject_codes = [str(subj_data[0]) for subj_data in subject_dict.values() if subj_data[0]]
        if subject_codes:
            interface_sheet["C6"].number_format = "@"
            dv_subject = DataValidation(type="list", formula1=f'"{",".join(subject_codes)}"', allow_blank=True)
            dv_subject.add("C6")
            interface_sheet.add_data_validation(dv_subject)
            logger.debug(f"Added dropdown to Interface!C6 with {len(subject_codes)} subject codes")
        
        # Save and close
        workbook.save(save_path)
        workbook.close()
        
        logger.info(f"Export completed successfully to {save_path}")
        return save_path
    
    except aiomysql.Error as e:
        logger.error(f"Error 1004: Database error: {e}")
        raise
    except Exception as e:
        logger.error(f"Error 1003: Excel operation failed: {e}")
        raise
    finally:
        if cursor:
            await cursor.close()
        if conn and pool:
            pool.release(conn)
        if pool:
            pool.close()
            await pool.wait_closed()


async def import_marks_from_excel_old(file_path: str) -> int:
    """Read marks from an Excel file and update student_subjects table using pandas. Return number of updated records."""
    try:
        if not os.path.exists(file_path):
            logger.error(f"Error 2001: Excel file not found at {file_path}")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Read Excel file using pandas
        df = pd.read_excel(file_path, sheet_name="Students", skiprows=1, usecols="A:N", engine="openpyxl")
        
        # Rename columns for clarity (based on your Excel structure)
        df.columns = [
            'student_id', 'full_name', 'sex', 'marks', 'subject_code', 'subject_name', 'subject_short',
            'centre_number', 'school_name', 'ward_name', 'council_name', 'region_name', 'exam_id', 'student_global_id'
        ]
        
        # Validate data
        logger.info(f"Read {len(df)} rows from {file_path}")
        df = df.dropna(subset=['exam_id', 'student_global_id', 'subject_code'])
        logger.info(f"After dropping missing keys, {len(df)} rows remain")
        
        # Convert marks to numeric, setting invalid values to None
        df['marks'] = pd.to_numeric(df['marks'], errors='coerce')
        
        # Create a DataFrame to group marks by exam_id, student_global_id, and clean subject_code
        df['clean_subject_code'] = df['subject_code'].apply(lambda x: x[:-2] if isinstance(x, str) and x.endswith("-P") else x)
        df['is_practical'] = df['subject_code'].where(df['subject_code'].notna(), '').astype(str).str.endswith("-P")

        df['theory_marks'] = df.apply(lambda row: row['marks'] if not row['is_practical'] else None, axis=1)
        df['practical_marks'] = df.apply(lambda row: row['marks'] if row['is_practical'] else None, axis=1)
        
        # Group by exam_id, student_global_id, clean_subject_code
        grouped = df.groupby(['exam_id', 'student_global_id', 'clean_subject_code']).agg({
            'theory_marks': 'first',
            'practical_marks': 'first'
        }).reset_index()
        
        logger.info(f"Grouped into {len(grouped)} unique exam_id, student_global_id, subject_code combinations")
        
        # Verify existing records in database
        pool = None
        conn = None
        cursor = None
        try:
            pool = await aiomysql.create_pool(**DB_CONFIG)
            conn = await pool.acquire()
            cursor = await conn.cursor()
            
            # Get unique exam_id for validation
            exam_ids = grouped['exam_id'].unique()
            if len(exam_ids) != 1:
                logger.error(f"Error 2004: Multiple exam_ids found in Excel: {exam_ids}")
                raise ValueError(f"Multiple exam_ids found in Excel: {exam_ids}")
            exam_id = exam_ids[0]
            
            await cursor.execute(
                "SELECT DISTINCT exam_id, student_global_id, subject_code FROM student_subjects WHERE exam_id = %s",
                (exam_id,)
            )
            existing_records = {(row[0], row[1], row[2]) for row in await cursor.fetchall()}
            
            # Filter out invalid records
            grouped = grouped[grouped.apply(
                lambda row: (row['exam_id'], row['student_global_id'], row['clean_subject_code']) in existing_records, axis=1
            )]
            logger.info(f"After filtering for existing records, {len(grouped)} rows remain")
            
            # Prepare SQLAlchemy engine
            engine = create_async_engine(
                f"mysql+aiomysql://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['db']}",
                pool_size=DB_CONFIG['maxsize']
            )
            
            async with AsyncSession(engine) as session:
                async with session.begin():
                    updated_records = 0
                    batch_size = 1000
                    for i in range(0, len(grouped), batch_size):
                        batch = grouped.iloc[i:i + batch_size]
                        query = """
                        UPDATE student_subjects
                        SET theory_marks = :theory_marks, practical_marks = :practical_marks
                        WHERE exam_id = :exam_id AND student_global_id = :student_global_id AND subject_code = :subject_code
                        """
                        params = [
                            {
                                'exam_id': row['exam_id'],
                                'student_global_id': row['student_global_id'],
                                'subject_code': row['clean_subject_code'],
                                'theory_marks': row['theory_marks'],
                                'practical_marks': row['practical_marks']
                            }
                            for _, row in batch.iterrows()
                        ]
                        
                        # Execute batch update
                        result = await session.execute(text(query), params)
                        batch_updated = result.rowcount
                        updated_records += batch_updated
                        logger.debug(f"Batch {i//batch_size + 1}: Updated {batch_updated} records")
                    
                    await session.commit()
                    logger.info(f"Successfully updated {updated_records} records from {file_path}")
                    return updated_records
        
        finally:
            if cursor:
                await cursor.close()
            if conn and pool:
                pool.release(conn)
            if pool:
                pool.close()
                await pool.wait_closed()
            if 'engine' in locals():
                await engine.dispose()
    
    except Exception as e:
        logger.error(f"Error 2003: Import operation failed: {e}")
        raise


async def import_marks_from_excel_OLDER(file_path: str) -> int:
    """Read marks from Excel, combine theory/practical records, clear rankings, update student_subjects table in bulk, and validate imports. Return number of updated records."""
    start_time = time.perf_counter()
    try:
        if not os.path.exists(file_path):
            logger.error(f"Error 2001: Excel file not found at {file_path}")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Read Excel file (columns A:N)
        df = pd.read_excel(file_path, sheet_name="Students", skiprows=1, usecols="A:N", engine="openpyxl")
        
        # Rename columns to match Excel structure
        df.columns = [
            'student_id', 'full_name', 'sex', 'marks', 'subject_code', 'subject_name', 'subject_short',
            'centre_number', 'school_name', 'ward_name', 'council_name', 'region_name', 'exam_id', 'student_global_id'
        ]
        
        # Validate data
        initial_count = len(df)
        df = df[['exam_id', 'student_global_id', 'subject_code', 'marks']].dropna(subset=['exam_id', 'student_global_id', 'subject_code'])
        logger.info(f"Read {initial_count} rows from {file_path}, {len(df)} remain after dropping missing keys")
        
        # Log sample data to verify column mapping
        if len(df) > 0:
            sample = df.head(2)
            for _, row in sample.iterrows():
                logger.debug(f"Sample row: exam_id={row['exam_id']}, student_global_id={row['student_global_id']}, subject_code={row['subject_code']}, marks={row['marks']}")
        
        # Validate single exam_id
        exam_ids = df['exam_id'].unique()
        if len(exam_ids) != 1:
            logger.error(f"Error 2004: Multiple exam_ids found in Excel: {exam_ids}")
            raise ValueError(f"Multiple exam_ids found in Excel: {exam_ids}")
        exam_id = exam_ids[0]
        
        # Convert marks to numeric
        df['marks'] = pd.to_numeric(df['marks'], errors='coerce')
        
        # Combine theory and practical marks
        df['clean_subject_code'] = df['subject_code'].apply(lambda x: x[:-2] if isinstance(x, str) and x.endswith("-P") else x)
        df['is_practical'] = df['subject_code'].where(df['subject_code'].notna(), '').astype(str).str.endswith("-P")
        df['theory_marks'] = df['marks'].where(~df['is_practical'])
        df['practical_marks'] = df['marks'].where(df['is_practical'])
        
        # Group by exam_id, student_global_id, clean_subject_code
        grouped = df.groupby(['exam_id', 'student_global_id', 'clean_subject_code']).agg({
            'theory_marks': 'first',
            'practical_marks': 'first'
        }).reset_index().rename(columns={'clean_subject_code': 'subject_code'})
        
        logger.info(f"Grouped into {len(grouped)} unique exam_id, student_global_id, subject_code combinations")
        
        # Log sample grouped data
        if len(grouped) > 0:
            sample = grouped.head(2)
            for _, row in sample.iterrows():
                logger.debug(f"Sample grouped: exam_id={row['exam_id']}, student_global_id={row['student_global_id']}, subject_code={row['subject_code']}, theory_marks={row['theory_marks']}, practical_marks={row['practical_marks']}")
        
        # Initialize database connection
        pool = None
        conn = None
        cursor = None
        try:
            pool = await aiomysql.create_pool(**DB_CONFIG)
            conn = await pool.acquire()
            cursor = await conn.cursor()
            
            # Check exam_id and clear rankings
            await cursor.execute("START TRANSACTION")
            await cursor.execute("SELECT COUNT(*) FROM student_subjects WHERE exam_id = %s", (exam_id,))
            exam_count = (await cursor.fetchone())[0]
            if exam_count == 0:
                logger.warning(f"No records found in student_subjects for exam_id={exam_id}. Updates will be skipped.")
                await cursor.execute("COMMIT")
                return 0
            
            # Log sample exam_id from database for debugging
            await cursor.execute("SELECT DISTINCT exam_id FROM student_subjects LIMIT 1")
            db_exam_id = await cursor.fetchone()
            if db_exam_id:
                logger.debug(f"Sample exam_id from student_subjects: {db_exam_id[0]}")
            
            # Count pre-import marks
            await cursor.execute(
                """
                SELECT COUNT(*) FROM student_subjects
                WHERE exam_id = %s AND (theory_marks >= 0 OR practical_marks >= 0)
                """,
                (exam_id,)
            )
            pre_import_count = (await cursor.fetchone())[0]
            logger.info(f"Pre-import: {pre_import_count} records have theory_marks >= 0 OR practical_marks >= 0 for exam_id={exam_id}")
            
            # Clear rankings
            clear_query = """
            UPDATE student_subjects
            SET overall_marks = NULL,
                subject_pos = NULL,
                ward_subject_pos = NULL,
                council_subject_pos = NULL,
                region_subject_pos = NULL,
                ward_subject_pos_gvt = NULL,
                ward_subject_pos_pvt = NULL,
                council_subject_pos_gvt = NULL,
                council_subject_pos_pvt = NULL,
                region_subject_pos_gvt = NULL,
                region_subject_pos_pvt = NULL,
                subject_out_of = NULL,
                ward_subject_out_of = NULL,
                council_subject_out_of = NULL,
                region_subject_out_of = NULL
            WHERE exam_id = %s
            """
            await cursor.execute(clear_query, (exam_id,))
            cleared_records = cursor.rowcount
            logger.info(f"Cleared rankings for {cleared_records} records for exam_id={exam_id}")
            
            # Fetch existing records efficiently
            await cursor.execute(
                """
                SELECT exam_id, student_global_id, subject_code
                FROM student_subjects
                WHERE exam_id = %s
                """,
                (exam_id,)
            )
            existing_records = {(row[0], row[1], row[2]) for row in await cursor.fetchall()}
            await cursor.execute("COMMIT")
            
            grouped = grouped[grouped.apply(
                lambda row: (row['exam_id'], row['student_global_id'], row['subject_code']) in existing_records, axis=1
            )]
            logger.info(f"After filtering for existing records, {len(grouped)} rows remain")
            
            # Bulk update with SQLAlchemy and retry logic
            engine = create_async_engine(
                f"mysql+aiomysql://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['db']}",
                pool_size=DB_CONFIG['maxsize']
            )
            
            async with AsyncSession(engine) as session:
                async with session.begin():
                    updated_records = 0
                    batch_size = 500
                    max_retries = 3
                    for i in range(0, len(grouped), batch_size):
                        batch = grouped.iloc[i:i + batch_size]
                        query = """
                        UPDATE student_subjects
                        SET theory_marks = :theory_marks, practical_marks = :practical_marks
                        WHERE exam_id = :exam_id AND student_global_id = :student_global_id AND subject_code = :subject_code
                        """
                        params = [
                            {
                                'exam_id': row['exam_id'],
                                'student_global_id': row['student_global_id'],
                                'subject_code': row['subject_code'],  # Use clean_subject_code
                                'theory_marks': row['theory_marks'] if pd.notnull(row['theory_marks']) else None,
                                'practical_marks': row['practical_marks'] if pd.notnull(row['practical_marks']) else None
                            }
                            for _, row in batch.iterrows()
                        ]
                        
                        # Retry logic for lock timeout
                        for attempt in range(max_retries):
                            try:
                                result = await session.execute(text(query), params)
                                batch_updated = result.rowcount
                                updated_records += batch_updated
                                logger.info(f"Batch {i//batch_size + 1}: Updated {batch_updated} records")
                                break
                            except aiomysql.OperationalError as e:
                                if '1205' in str(e):  # Lock wait timeout
                                    if attempt < max_retries - 1:
                                        logger.warning(f"Lock wait timeout on batch {i//batch_size + 1}, attempt {attempt + 1}. Retrying in 5 seconds...")
                                        await asyncio.sleep(5)
                                        continue
                                    logger.error(f"Error 2005: Max retries exceeded for batch {i//batch_size + 1}: {e}")
                                    raise
                                raise
                    
                    # Validate imported marks
                    await cursor.execute(
                        """
                        SELECT COUNT(*) FROM student_subjects
                        WHERE exam_id = %s AND (theory_marks >= 0 OR practical_marks >= 0)
                        """,
                        (exam_id,)
                    )
                    post_import_count = (await cursor.fetchone())[0]
                    logger.info(f"Post-import: {post_import_count} records have theory_marks >= 0 OR practical_marks >= 0 for exam_id={exam_id}")
                    if post_import_count > pre_import_count:
                        logger.info(f"Imported {post_import_count - pre_import_count} new marks")
                    elif post_import_count == pre_import_count:
                        logger.warning("No new marks imported (counts unchanged)")
                    else:
                        logger.warning(f"Marks reduced from {pre_import_count} to {post_import_count}, possible data loss")
                    
                    await session.commit()
                    end_time = time.perf_counter()
                    elapsed_time = end_time - start_time
                    minutes, seconds = divmod(elapsed_time, 60)
                    logger.info(f"Successfully updated {updated_records} records from {file_path} in {int(minutes)} minutes, {seconds:.2f} seconds")
                    return updated_records
        
        finally:
            if cursor:
                await cursor.close()
            if conn and pool:
                pool.release(conn)
            if pool:
                pool.close()
                await pool.wait_closed()
            if 'engine' in locals():
                await engine.dispose()
    
    except Exception as e:
        logger.error(f"Error 2003: Import operation failed: {e}")
        raise


async def calculate_marks_and_ranks(exam_id: str) -> int:
    """Calculate overall_marks and rankings for student_subjects table using pandas. Return number of updated records."""
    info=await clear_student_subject_results(exam_id)
    print(info)
    start_time = time.perf_counter()
    pool = None
    conn = None
    cursor = None
    
    try:
        pool = await aiomysql.create_pool(**DB_CONFIG)
        conn = await pool.acquire()
        cursor = await conn.cursor(aiomysql.DictCursor)
        
        # Fetch records where theory_marks >= 0 OR practical_marks >= 0
        query = """
        SELECT ss.exam_id, ss.student_global_id, ss.subject_code, 
               ss.theory_marks, ss.practical_marks,
               es.has_practical,
               s.ward_name, s.council_name, s.region_name, s.school_type
        FROM student_subjects ss
        INNER JOIN exam_subjects es ON ss.exam_id = es.exam_id AND ss.subject_code = es.subject_code
        INNER JOIN students st ON ss.student_global_id = st.student_global_id AND ss.exam_id = st.exam_id
        INNER JOIN schools s ON ss.centre_number = s.centre_number
        WHERE ss.exam_id = %s AND (ss.theory_marks >= 0 OR ss.practical_marks >= 0)
        """
        await cursor.execute(query, (exam_id,))
        records = await cursor.fetchall()
        
        if not records:
            logger.warning(f"No records found for exam_id={exam_id}")
            return 0
        
        # Convert to DataFrame
        df = pd.DataFrame(records)
        logger.info(f"Found {len(df)} Entries in exam {exam_id}")
        
        # Calculate overall_marks
        df['overall_marks'] = None
        has_practical = df['has_practical'] == True
        both_marks = (df['theory_marks'].notnull() & df['practical_marks'].notnull())
        theory_only = (df['theory_marks'].notnull() & df['practical_marks'].isnull())
        practical_only = (df['theory_marks'].isnull() & df['practical_marks'].notnull())
        
        df.loc[has_practical & both_marks, 'overall_marks'] = (df['theory_marks'] + df['practical_marks']) * 2 / 3
        df.loc[has_practical & theory_only, 'overall_marks'] = df['theory_marks'] * 2 / 3
        df.loc[has_practical & practical_only, 'overall_marks'] = df['practical_marks'] * 2 / 3
        df.loc[~has_practical, 'overall_marks'] = df['theory_marks']
        
        # Log sample records (first 5 and last 5 if >10, else all)
        sample_indices = list(df.index[:5]) + list(df.index[-5:]) if len(df) > 10 else df.index
        for i in sample_indices:
            row = df.iloc[i]
            logger.info(f"  Processing {row['student_global_id']} - {row['subject_code']} - {row['theory_marks']} - {row['practical_marks']} | {i+1} out of {len(df)}")
        
        # Normalize school_type
        df['school_type'] = df['school_type'].apply(lambda x: x if x in ['GOVERNMENT', 'PRIVATE'] else 'UNKNOWN')
        
        # Initialize ranking and out_of columns
        rank_cols = ['subject_pos', 'ward_subject_pos', 'council_subject_pos', 'region_subject_pos',
                     'ward_subject_pos_gvt', 'ward_subject_pos_pvt', 'council_subject_pos_gvt',
                     'council_subject_pos_pvt', 'region_subject_pos_gvt', 'region_subject_pos_pvt']
        out_of_cols = ['subject_out_of', 'ward_subject_out_of', 'council_subject_out_of', 'region_subject_out_of']
        df[rank_cols + out_of_cols] = None
        
        # Compute rankings and out_of counts
        def compute_ranks(df, group_cols, pos_col, out_of_col, valid_mask=None):
            temp_df = df[df['overall_marks'].notnull()] if valid_mask is None else df[valid_mask & df['overall_marks'].notnull()]
            if not temp_df.empty:
                df.loc[temp_df.index, pos_col] = temp_df.groupby(group_cols)['overall_marks'].rank(method='dense', ascending=False).astype('Int64')
                df.loc[temp_df.index, out_of_col] = temp_df.groupby(group_cols)['overall_marks'].transform('count').astype('Int64')
        
        # Overall position
        compute_ranks(df, ['subject_code'], 'subject_pos', 'subject_out_of')
        
        # Ward positions
        compute_ranks(df, ['subject_code', 'ward_name'], 'ward_subject_pos', 'ward_subject_out_of',
                      df['ward_name'].notnull())
        compute_ranks(df, ['subject_code', 'ward_name'], 'ward_subject_pos_gvt', 'ward_subject_out_of',
                      (df['ward_name'].notnull() & (df['school_type'] == 'GOVERNMENT')))
        compute_ranks(df, ['subject_code', 'ward_name'], 'ward_subject_pos_pvt', 'ward_subject_out_of',
                      (df['ward_name'].notnull() & (df['school_type'] == 'PRIVATE')))
        
        # Council positions
        compute_ranks(df, ['subject_code', 'council_name'], 'council_subject_pos', 'council_subject_out_of',
                      df['council_name'].notnull())
        compute_ranks(df, ['subject_code', 'council_name'], 'council_subject_pos_gvt', 'council_subject_out_of',
                      (df['council_name'].notnull() & (df['school_type'] == 'GOVERNMENT')))
        compute_ranks(df, ['subject_code', 'council_name'], 'council_subject_pos_pvt', 'council_subject_out_of',
                      (df['council_name'].notnull() & (df['school_type'] == 'PRIVATE')))
        
        # Region positions
        compute_ranks(df, ['subject_code', 'region_name'], 'region_subject_pos', 'region_subject_out_of',
                      df['region_name'].notnull())
        compute_ranks(df, ['subject_code', 'region_name'], 'region_subject_pos_gvt', 'region_subject_out_of',
                      (df['region_name'].notnull() & (df['school_type'] == 'GOVERNMENT')))
        compute_ranks(df, ['subject_code', 'region_name'], 'region_subject_pos_pvt', 'region_subject_out_of',
                      (df['region_name'].notnull() & (df['school_type'] == 'PRIVATE')))
        
        # Set gvt/pvt positions to NULL for UNKNOWN school_type
        df.loc[df['school_type'] == 'UNKNOWN', ['ward_subject_pos_gvt', 'ward_subject_pos_pvt',
                                               'council_subject_pos_gvt', 'council_subject_pos_pvt',
                                               'region_subject_pos_gvt', 'region_subject_pos_pvt']] = None
        
        # Prepare SQLAlchemy engine
        engine = create_async_engine(
            f"mysql+aiomysql://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['db']}",
            pool_size=DB_CONFIG['maxsize']
        )
        
        # Bulk update database
        logger.info(f"Starting Database Update==========")
        async with AsyncSession(engine) as session:
            async with session.begin():
                updated_records = 0
                batch_size = 1000
                for i in range(0, len(df), batch_size):
                    batch = df.iloc[i:i + batch_size]
                    query = """
                    UPDATE student_subjects
                    SET overall_marks = :overall_marks,
                        subject_pos = :subject_pos,
                        ward_subject_pos = :ward_subject_pos,
                        council_subject_pos = :council_subject_pos,
                        region_subject_pos = :region_subject_pos,
                        ward_subject_pos_gvt = :ward_subject_pos_gvt,
                        ward_subject_pos_pvt = :ward_subject_pos_pvt,
                        council_subject_pos_gvt = :council_subject_pos_gvt,
                        council_subject_pos_pvt = :council_subject_pos_pvt,
                        region_subject_pos_gvt = :region_subject_pos_gvt,
                        region_subject_pos_pvt = :region_subject_pos_pvt,
                        subject_out_of = :subject_out_of,
                        ward_subject_out_of = :ward_subject_out_of,
                        council_subject_out_of = :council_subject_out_of,
                        region_subject_out_of = :region_subject_out_of
                    WHERE exam_id = :exam_id AND student_global_id = :student_global_id AND subject_code = :subject_code
                    """
                    params = [
                        {
                            'exam_id': row['exam_id'],
                            'student_global_id': row['student_global_id'],
                            'subject_code': row['subject_code'],
                            'overall_marks': row['overall_marks'] if pd.notnull(row['overall_marks']) else None,
                            'subject_pos': row['subject_pos'] if pd.notnull(row['subject_pos']) else None,
                            'ward_subject_pos': row['ward_subject_pos'] if pd.notnull(row['ward_subject_pos']) else None,
                            'council_subject_pos': row['council_subject_pos'] if pd.notnull(row['council_subject_pos']) else None,
                            'region_subject_pos': row['region_subject_pos'] if pd.notnull(row['region_subject_pos']) else None,
                            'ward_subject_pos_gvt': row['ward_subject_pos_gvt'] if pd.notnull(row['ward_subject_pos_gvt']) else None,
                            'ward_subject_pos_pvt': row['ward_subject_pos_pvt'] if pd.notnull(row['ward_subject_pos_pvt']) else None,
                            'council_subject_pos_gvt': row['council_subject_pos_gvt'] if pd.notnull(row['council_subject_pos_gvt']) else None,
                            'council_subject_pos_pvt': row['council_subject_pos_pvt'] if pd.notnull(row['council_subject_pos_pvt']) else None,
                            'region_subject_pos_gvt': row['region_subject_pos_gvt'] if pd.notnull(row['region_subject_pos_gvt']) else None,
                            'region_subject_pos_pvt': row['region_subject_pos_pvt'] if pd.notnull(row['region_subject_pos_pvt']) else None,
                            'subject_out_of': row['subject_out_of'] if pd.notnull(row['subject_out_of']) else None,
                            'ward_subject_out_of': row['ward_subject_out_of'] if pd.notnull(row['ward_subject_out_of']) else None,
                            'council_subject_out_of': row['council_subject_out_of'] if pd.notnull(row['council_subject_out_of']) else None,
                            'region_subject_out_of': row['region_subject_out_of'] if pd.notnull(row['region_subject_out_of']) else None
                        }
                        for _, row in batch.iterrows()
                    ]
                    
                    result = await session.execute(text(query), params)
                    batch_updated = result.rowcount
                    updated_records += batch_updated
                    logger.info(f"Batch {i//batch_size + 1}: Updated {batch_updated} records")
        
        await session.commit()
        end_time = time.perf_counter()
        elapsed_time = end_time - start_time
        minutes, seconds = divmod(elapsed_time, 60)
        logger.info(f"Successfully updated {updated_records} records for exam_id={exam_id} in {int(minutes)} minutes, {seconds:.2f} seconds")
        
        return updated_records
    
    except aiomysql.Error as e:
        logger.error(f"Error 3001: Database error: {e}")
        if conn:
            await conn.rollback()
        raise
    except Exception as e:
        logger.error(f"Error 3002: Operation failed: {e}")
        if conn:
            await conn.rollback()
        raise
    finally:
        if cursor:
            await cursor.close()
        if conn and pool:
            pool.release(conn)
        if pool:
            pool.close()
            await pool.wait_closed()
        if 'engine' in locals():
            await engine.dispose()


async def process_exam_results_OLD(exam_id: str):
    logger.info(f"Starting exam results processing for exam_id: {exam_id}")
    
    # Create aiomysql connection pool
    logger.info(f"Creating aiomysql connection pool for database: {DB_CONFIG['db']}")
    pool = await aiomysql.create_pool(
        host=DB_CONFIG['host'],
        port=DB_CONFIG['port'],
        user=DB_CONFIG['user'],
        password=DB_CONFIG['password'],
        db=DB_CONFIG['db'],
        maxsize=DB_CONFIG['maxsize'],
        charset='utf8mb4'
    )

    async with pool.acquire() as conn:
        async with conn.cursor(aiomysql.DictCursor) as cur:
            try:
                # Load reference data
                logger.info(f"Loading exam grades for exam_id: {exam_id}")
                grades_query = "SELECT exam_id, grade, lower_value, highest_value, grade_points, division_points FROM exam_grades WHERE exam_id = %s"
                await cur.execute(grades_query, (exam_id,))
                grades_df = pd.DataFrame(await cur.fetchall(), columns=['exam_id', 'grade', 'lower_value', 'highest_value', 'grade_points', 'division_points'])
                logger.info(f"Loaded {len(grades_df)} grade records")

                logger.info(f"Loading exam divisions for exam_id: {exam_id}")
                divisions_query = "SELECT exam_id, division, lowest_points, highest_points FROM exam_divisions WHERE exam_id = %s"
                await cur.execute(divisions_query, (exam_id,))
                divisions_df = pd.DataFrame(await cur.fetchall(), columns=['exam_id', 'division', 'lowest_points', 'highest_points'])
                logger.info(f"Loaded {len(divisions_df)} division records")

                logger.info(f"Loading exam data for exam_id: {exam_id}")
                exams_query = "SELECT exam_id, avg_style FROM exams WHERE exam_id = %s"
                await cur.execute(exams_query, (exam_id,))
                exams_df = pd.DataFrame(await cur.fetchall(), columns=['exam_id', 'avg_style'])
                logger.info(f"Loaded {len(exams_df)} exam records")

                if exams_df.empty:
                    logger.error(f"No exam found for exam_id: {exam_id}")
                    return

                avg_style = exams_df['avg_style'].iloc[0]
                logger.info(f"Exam avg_style: {avg_style}")

                logger.info(f"Loading schools with at least one student record")
                schools_query = """
                    SELECT DISTINCT s.centre_number, s.region_name, s.council_name, s.ward_name, s.school_type
                    FROM schools s
                    INNER JOIN students st ON s.centre_number = st.centre_number
                """
                await cur.execute(schools_query)
                schools_df = pd.DataFrame(await cur.fetchall(), columns=['centre_number', 'region_name', 'council_name', 'ward_name', 'school_type'])
                logger.info(f"Loaded {len(schools_df)} school records with student data")

                # Load student subjects
                logger.info(f"Loading student subjects for exam_id: {exam_id}")
                subjects_query = """
                    SELECT id, exam_id, student_global_id, centre_number, overall_marks, subject_code
                    FROM student_subjects WHERE exam_id = %s
                """
                await cur.execute(subjects_query, (exam_id,))
                subjects_df = pd.DataFrame(await cur.fetchall(), columns=['id', 'exam_id', 'student_global_id', 'centre_number', 'overall_marks', 'subject_code'])
                logger.info(f"Loaded {len(subjects_df)} student subject records")

                if subjects_df.empty:
                    logger.warning(f"No student subjects found for exam_id: {exam_id}")
                    return

                # Map grades
                logger.info(f"Mapping grades for {len(subjects_df)} records")
                def map_grade(marks):
                    if pd.notnull(marks):
                        grade_row = grades_df[(grades_df['lower_value'] < marks) & (marks <= grades_df['highest_value'])]
                        return grade_row['grade'].iloc[0] if not grade_row.empty else None
                    return None

                def map_grade_points(marks):
                    if pd.notnull(marks):
                        grade_row = grades_df[(grades_df['lower_value'] < marks) & (marks <= grades_df['highest_value'])]
                        return grade_row['grade_points'].iloc[0] if not grade_row.empty else None
                    return None

                subjects_df['subject_grade'] = subjects_df['overall_marks'].apply(map_grade)
                subjects_df['grade_points'] = subjects_df['overall_marks'].apply(map_grade_points)
                logger.info(f"Completed grade mapping")

                # Update student_subjects
                logger.info(f"Updating student_subjects for {len(subjects_df)} records")
                for i, row in subjects_df.iterrows():
                    await cur.execute(
                        "UPDATE student_subjects SET subject_grade = %s WHERE id = %s",
                        (row['subject_grade'], row['id'])
                    )
                    if (i + 1) % 1000 == 0:
                        logger.info(f"Updated {i + 1} student subject records")
                        await conn.commit()
                await conn.commit()
                logger.info(f"Finished updating student_subjects")

                # Compute results
                logger.info(f"Computing results for exam_id: {exam_id}")
                results = subjects_df.groupby(['exam_id', 'student_global_id', 'centre_number']).agg({
                    'overall_marks': ['sum', 'count', lambda x: x.nlargest(7).sum()],
                    'grade_points': lambda x: x.nlargest(7).sum()
                }).reset_index()
                results.columns = ['exam_id', 'student_global_id', 'centre_number', 'total_marks', 'subject_count', 'best_7_marks', 'total_points']
                logger.info(f"Computed results for {len(results)} students")

                # Compute avg_marks
                logger.info(f"Computing average marks with style: {avg_style}")
                if avg_style == 'AUTO':
                    results['avg_marks'] = np.where(
                        results['subject_count'] >= 7,
                        results['total_marks'] / results['subject_count'],
                        results['total_marks'] / 7
                    )
                elif avg_style == 'SEVEN_BEST':
                    results['avg_marks'] = results['best_7_marks'] / 7
                elif avg_style == 'EIGHT_BEST':
                    results['avg_marks'] = results['best_7_marks'] / 8
                logger.info(f"Completed average marks calculation")

                # Map avg_grade
                logger.info(f"Mapping average grades for {len(results)} records")
                results['avg_grade'] = results['avg_marks'].apply(map_grade)
                logger.info(f"Completed average grade mapping")

                # Map division
                logger.info(f"Mapping divisions for {len(results)} records")
                results['division'] = results.apply(
                    lambda row: divisions_df[
                        (divisions_df['lowest_points'] <= row['total_points']) & 
                        (row['total_points'] <= divisions_df['highest_points'])
                    ]['division'].iloc[0] if pd.notnull(row['total_points']) and not divisions_df[
                        (divisions_df['lowest_points'] <= row['total_points']) & 
                        (row['total_points'] <= divisions_df['highest_points'])
                    ].empty else 'INC' if row['subject_count'] < 7 else 'ABS',
                    axis=1
                )
                logger.info(f"Completed division mapping")

                # Merge school data
                logger.info(f"Merging school data")
                results = results.merge(schools_df, on='centre_number', how='left')
                logger.info(f"Completed school data merge")

                # Initialize ranking columns to avoid index errors
                logger.info(f"Initializing ranking columns")
                for col in ['ward_pos', 'ward_out_of', 'council_pos', 'council_out_of', 'region_pos', 'region_out_of',
                            'ward_pos_gvt', 'ward_pos_pvt', 'council_pos_gvt', 'council_pos_pvt', 'region_pos_gvt', 'region_pos_pvt']:
                    results[col] = pd.NA
                results = results.astype({col: 'Int64' for col in ['ward_pos', 'ward_out_of', 'council_pos', 'council_out_of', 
                                                                  'region_pos', 'region_out_of', 'ward_pos_gvt', 'ward_pos_pvt', 
                                                                  'council_pos_gvt', 'council_pos_pvt', 'region_pos_gvt', 'region_pos_pvt']})
                logger.info(f"Initialized ranking columns")

                # Compute rankings
                logger.info(f"Computing rankings for exam_id: {exam_id}")
                results['pos'] = results[results['avg_marks'].notnull()]['avg_marks'].rank(ascending=False, method='dense').astype('Int64')
                results['out_of'] = results[results['avg_marks'].notnull()].shape[0]
                logger.info(f"Computed overall rankings: {results['out_of'].iloc[0]} students")

                for level in ['ward', 'council', 'region']:
                    logger.info(f"Computing {level} rankings")
                    for school_type in ['GOVERNMENT', 'PRIVATE']:
                        for name in results[f'{level}_name'].dropna().unique():
                            mask = (results[f'{level}_name'] == name) & (results['school_type'] == school_type) & (results['avg_marks'].notnull())
                            if mask.any():
                                results.loc[mask, f'{level}_pos_{school_type[:3].lower()}'] = results[mask]['avg_marks'].rank(ascending=False, method='dense').astype('Int64')
                            mask_all = (results[f'{level}_name'] == name) & (results['avg_marks'].notnull())
                            if mask_all.any():
                                results.loc[mask_all, f'{level}_pos'] = results[mask_all]['avg_marks'].rank(ascending=False, method='dense').astype('Int64')
                                results.loc[mask_all, f'{level}_out_of'] = results[mask_all].shape[0]
                    logger.info(f"Completed {level} rankings")

                # Prepare results for insertion
                logger.info(f"Preparing {len(results)} results for insertion")
                results['id'] = [str(uuid6()) for _ in range(len(results))]
                results['created_at'] = pd.Timestamp.now()
                results = results[['id', 'exam_id', 'student_global_id', 'centre_number', 'avg_marks', 'avg_grade', 
                                 'total_marks', 'division', 'total_points', 'pos', 'out_of', 
                                 'ward_pos', 'ward_out_of', 'council_pos', 'council_out_of', 
                                 'region_pos', 'region_out_of', 'ward_pos_gvt', 'ward_pos_pvt', 
                                 'council_pos_gvt', 'council_pos_pvt', 'region_pos_gvt', 'region_pos_pvt', 
                                 'created_at']]
                logger.info(f"Prepared results DataFrame")

                # Insert results
                logger.info(f"Inserting {len(results)} results into results table")
                for i, row in results.iterrows():
                    await cur.execute(
                        """
                        INSERT INTO results (id, exam_id, student_global_id, centre_number, avg_marks, avg_grade, 
                                            total_marks, division, total_points, pos, out_of, 
                                            ward_pos, ward_out_of, council_pos, council_out_of, 
                                            region_pos, region_out_of, ward_pos_gvt, ward_pos_pvt, 
                                            council_pos_gvt, council_pos_pvt, region_pos_gvt, region_pos_pvt, 
                                            created_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        tuple(row)
                    )
                    if (i + 1) % 1000 == 0:
                        logger.info(f"Inserted {i + 1} result records")
                        await conn.commit()
                await conn.commit()
                logger.info(f"Completed insertion of {len(results)} results")

                logger.info(f"Processing completed successfully for exam_id: {exam_id}")
        
            except Exception as e:
                logger.error(f"Error during processing for exam_id {exam_id}: {str(e)}")
                await conn.rollback()
                raise
            finally:
                logger.info(f"Closing connection pool")
                pool.close()
                await pool.wait_closed()
                logger.info(f"Connection pool closed")


async def process_exam_results(exam_id: str):
    pool = None
    try:
        # Create aiomysql connection pool
        logger.info(f"Creating aiomysql connection pool for database: {DB_CONFIG['db']}")
        pool = await aiomysql.create_pool(
            host=DB_CONFIG['host'],
            port=DB_CONFIG['port'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            db=DB_CONFIG['db'],
            maxsize=DB_CONFIG['maxsize'],
            charset='utf8mb4'
        )

        async with pool.acquire() as conn:
            async with conn.cursor(aiomysql.DictCursor) as cur:
                # Load reference data
                logger.info(f"Loading exam grades for exam_id: {exam_id}")
                grades_query = "SELECT exam_id, grade, lower_value, highest_value, grade_points, division_points FROM exam_grades WHERE exam_id = %s"
                await cur.execute(grades_query, (exam_id,))
                grades_df = pd.DataFrame(await cur.fetchall(), columns=['exam_id', 'grade', 'lower_value', 'highest_value', 'grade_points', 'division_points'])
                logger.info(f"Loaded {len(grades_df)} grade records. dtypes: {grades_df.dtypes.to_dict()}")

                logger.info(f"Loading exam divisions for exam_id: {exam_id}")
                divisions_query = "SELECT exam_id, division, lowest_points, highest_points FROM exam_divisions WHERE exam_id = %s"
                await cur.execute(divisions_query, (exam_id,))
                divisions_df = pd.DataFrame(await cur.fetchall(), columns=['exam_id', 'division', 'lowest_points', 'highest_points'])
                logger.info(f"Loaded {len(divisions_df)} division records. dtypes: {divisions_df.dtypes.to_dict()}")

                logger.info(f"Loading exam data for exam_id: {exam_id}")
                exams_query = "SELECT exam_id, avg_style FROM exams WHERE exam_id = %s"
                await cur.execute(exams_query, (exam_id,))
                exams_df = pd.DataFrame(await cur.fetchall(), columns=['exam_id', 'avg_style'])
                logger.info(f"Loaded {len(exams_df)} exam records. dtypes: {exams_df.dtypes.to_dict()}")

                if exams_df.empty:
                    logger.error(f"No exam found for exam_id: {exam_id}")
                    return

                avg_style = exams_df['avg_style'].iloc[0]
                logger.info(f"Exam avg_style: {avg_style}")

                logger.info(f"Loading schools with at least one student record")
                schools_query = """
                    SELECT DISTINCT s.centre_number, s.region_name, s.council_name, s.ward_name, s.school_type
                    FROM schools s
                    INNER JOIN students st ON s.centre_number = st.centre_number
                """
                await cur.execute(schools_query)
                schools_df = pd.DataFrame(await cur.fetchall(), columns=['centre_number', 'region_name', 'council_name', 'ward_name', 'school_type'])
                logger.info(f"Loaded {len(schools_df)} school records with student data. dtypes: {schools_df.dtypes.to_dict()}")

                # Load student subjects
                logger.info(f"Loading student subjects for exam_id: {exam_id}")
                subjects_query = """
                    SELECT id, exam_id, student_global_id, centre_number, overall_marks, subject_code
                    FROM student_subjects WHERE exam_id = %s
                """
                await cur.execute(subjects_query, (exam_id,))
                subjects_df = pd.DataFrame(await cur.fetchall(), columns=['id', 'exam_id', 'student_global_id', 'centre_number', 'overall_marks', 'subject_code'])
                logger.info(f"Loaded {len(subjects_df)} student subject records. dtypes: {subjects_df.dtypes.to_dict()}")

                if subjects_df.empty:
                    logger.warning(f"No student subjects found for exam_id: {exam_id}")
                    return

                # Ensure numeric types
                logger.info("Converting overall_marks to float64")
                subjects_df['overall_marks'] = pd.to_numeric(subjects_df['overall_marks'], errors='coerce').astype('float64')
                logger.info(f"overall_marks dtypes after conversion: {subjects_df['overall_marks'].dtype}")
                logger.info(f"Non-numeric overall_marks values: {subjects_df['overall_marks'].isna().sum()} NaN values")

                # Map grades
                logger.info(f"Mapping grades for {len(subjects_df)} records")
                def map_grade(marks):
                    if pd.notnull(marks):
                        grade_row = grades_df[(grades_df['lower_value'] < marks) & (marks <= grades_df['highest_value'])]
                        return grade_row['grade'].iloc[0] if not grade_row.empty else None
                    return None

                def map_grade_points(marks):
                    if pd.notnull(marks):
                        grade_row = grades_df[(grades_df['lower_value'] < marks) & (marks <= grades_df['highest_value'])]
                        return float(grade_row['grade_points'].iloc[0]) if not grade_row.empty else np.nan
                    return np.nan

                subjects_df['subject_grade'] = subjects_df['overall_marks'].apply(map_grade)
                subjects_df['grade_points'] = subjects_df['overall_marks'].apply(map_grade_points).astype('float64')
                logger.info(f"Completed grade mapping. subjects_df dtypes: {subjects_df.dtypes.to_dict()}")

                # Update student_subjects
                logger.info(f"Updating student_subjects for {len(subjects_df)} records")
                for i, row in subjects_df.iterrows():
                    await cur.execute(
                        "UPDATE student_subjects SET subject_grade = %s WHERE id = %s",
                        (row['subject_grade'], row['id'])
                    )
                    if (i + 1) % 1000 == 0:
                        logger.info(f"Updated {i + 1} student subject records")
                        await conn.commit()
                await conn.commit()
                logger.info(f"Finished updating student_subjects")

                # Compute results
                logger.info(f"Computing results for exam_id: {exam_id}")
                results = subjects_df.groupby(['exam_id', 'student_global_id', 'centre_number']).agg({
                    'overall_marks': ['sum', 'count', lambda x: x.nlargest(7).sum()],
                    'grade_points': lambda x: x.nlargest(7).sum()
                }).reset_index()
                results.columns = ['exam_id', 'student_global_id', 'centre_number', 'total_marks', 'subject_count', 'best_7_marks', 'total_points']
                logger.info(f"Computed results for {len(results)} students. dtypes: {results.dtypes.to_dict()}")

                # Compute avg_marks
                logger.info(f"Computing average marks with style: {avg_style}")
                if avg_style == 'AUTO':
                    results['avg_marks'] = np.where(
                        results['subject_count'] >= 7,
                        results['total_marks'] / results['subject_count'],
                        results['total_marks'] / 7
                    )
                elif avg_style == 'SEVEN_BEST':
                    results['avg_marks'] = results['best_7_marks'] / 7
                elif avg_style == 'EIGHT_BEST':
                    results['avg_marks'] = results['best_7_marks'] / 8
                logger.info(f"Completed average marks calculation. avg_marks dtype: {results['avg_marks'].dtype}")

                # Map avg_grade
                logger.info(f"Mapping average grades for {len(results)} records")
                results['avg_grade'] = results['avg_marks'].apply(map_grade)
                logger.info(f"Completed average grade mapping. avg_grade dtype: {results['avg_grade'].dtype}")

                # Map division
                logger.info(f"Mapping divisions for {len(results)} records")
                results['division'] = results.apply(
                    lambda row: divisions_df[
                        (divisions_df['lowest_points'] <= row['total_points']) & 
                        (row['total_points'] <= divisions_df['highest_points'])
                    ]['division'].iloc[0] if pd.notnull(row['total_points']) and not divisions_df[
                        (divisions_df['lowest_points'] <= row['total_points']) & 
                        (row['total_points'] <= divisions_df['highest_points'])
                    ].empty else 'INC' if row['subject_count'] < 7 else 'ABS',
                    axis=1
                )
                logger.info(f"Completed division mapping. division dtype: {results['division'].dtype}")

                # Merge school data
                logger.info(f"Merging school data")
                results = results.merge(schools_df, on='centre_number', how='left')
                logger.info(f"Completed school data merge. dtypes: {results.dtypes.to_dict()}")

                # Initialize ranking columns
                logger.info(f"Initializing ranking columns")
                for col in ['ward_pos', 'ward_out_of', 'council_pos', 'council_out_of', 'region_pos', 'region_out_of',
                            'ward_pos_gvt', 'ward_pos_pvt', 'council_pos_gvt', 'council_pos_pvt', 'region_pos_gvt', 'region_pos_pvt']:
                    results[col] = pd.NA
                results = results.astype({col: 'Int64' for col in ['ward_pos', 'ward_out_of', 'council_pos', 'council_out_of', 
                                                                  'region_pos', 'region_out_of', 'ward_pos_gvt', 'ward_pos_pvt', 
                                                                  'council_pos_gvt', 'council_pos_pvt', 'region_pos_gvt', 'region_pos_pvt']})
                logger.info(f"Initialized ranking columns. dtypes: {results.dtypes.to_dict()}")

                # Compute rankings
                logger.info(f"Computing rankings for exam_id: {exam_id}")
                results['pos'] = results[results['avg_marks'].notnull()]['avg_marks'].rank(ascending=False, method='dense').astype('Int64')
                results['out_of'] = results[results['avg_marks'].notnull()].shape[0]
                logger.info(f"Computed overall rankings: {results['out_of'].iloc[0]} students. pos dtype: {results['pos'].dtype}")

                for level in ['ward', 'council', 'region']:
                    logger.info(f"Computing {level} rankings")
                    for school_type in ['GOVERNMENT', 'PRIVATE']:
                        for name in results[f'{level}_name'].dropna().unique():
                            mask = (results[f'{level}_name'] == name) & (results['school_type'] == school_type) & (results['avg_marks'].notnull())
                            if mask.any():
                                results.loc[mask, f'{level}_pos_{school_type[:3].lower()}'] = results[mask]['avg_marks'].rank(ascending=False, method='dense').astype('Int64')
                            mask_all = (results[f'{level}_name'] == name) & (results['avg_marks'].notnull())
                            if mask_all.any():
                                results.loc[mask_all, f'{level}_pos'] = results[mask_all]['avg_marks'].rank(ascending=False, method='dense').astype('Int64')
                                results.loc[mask_all, f'{level}_out_of'] = results[mask_all].shape[0]
                    logger.info(f"Completed {level} rankings. {level}_pos dtype: {results[f'{level}_pos'].dtype}")

                # Prepare results for insertion
                logger.info(f"Preparing {len(results)} results for insertion")
                results['id'] = [str(uuid6()) for _ in range(len(results))]
                results['created_at'] = pd.Timestamp.now()
                results = results[['id', 'exam_id', 'student_global_id', 'centre_number', 'avg_marks', 'avg_grade', 
                                 'total_marks', 'division', 'total_points', 'pos', 'out_of', 
                                 'ward_pos', 'ward_out_of', 'council_pos', 'council_out_of', 
                                 'region_pos', 'region_out_of', 'ward_pos_gvt', 'ward_pos_pvt', 
                                 'council_pos_gvt', 'council_pos_pvt', 'region_pos_gvt', 'region_pos_pvt', 
                                 'created_at']]
                logger.info(f"Prepared results DataFrame. dtypes: {results.dtypes.to_dict()}")

                # Insert results
                logger.info(f"Inserting {len(results)} results into results table")
                for i, row in results.iterrows():
                    await cur.execute(
                        """
                        INSERT INTO results (id, exam_id, student_global_id, centre_number, avg_marks, avg_grade, 
                                            total_marks, division, total_points, pos, out_of, 
                                            ward_pos, ward_out_of, council_pos, council_out_of, 
                                            region_pos, region_out_of, ward_pos_gvt, ward_pos_pvt, 
                                            council_pos_gvt, council_pos_pvt, region_pos_gvt, region_pos_pvt, 
                                            created_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        tuple(row)
                    )
                    if (i + 1) % 1000 == 0:
                        logger.info(f"Inserted {i + 1} result records")
                        await conn.commit()
                await conn.commit()
                logger.info(f"Completed insertion of {len(results)} results")

                logger.info(f"Processing completed successfully for exam_id: {exam_id}")

    except Exception as e:
        logger.error(f"Error during processing for exam_id {exam_id}: {str(e)}")
        await conn.rollback()
        raise
    finally:
        logger.info("Closing connection pool")
        if pool is not None:
            pool.close()
            await pool.wait_closed()
        logger.info("Connection pool closed")


async def clear_student_subject_results(exam_id: str, clear_overall_marks: bool = False) -> Dict[str, Any]:
    """
    Clears results for a specific exam in the StudentSubject table using aiomysql.
    
    Args:
        exam_id (str): The exam ID to clear results for
        clear_overall_marks (bool): Whether to also clear overall_marks (default False)
    
    Returns:
        Dict[str, Any]: Status dictionary with 'status' and 'message' keys
    """
    # Base SQL statement
    sql = """
        UPDATE student_subjects
        SET 
            theory_marks = NULL,
            practical_marks = NULL,
            subject_grade = NULL,
            subject_pos = NULL,
            subject_out_of = NULL,
            ward_subject_pos = NULL,
            ward_subject_out_of = NULL,
            council_subject_pos = NULL,
            council_subject_out_of = NULL,
            region_subject_pos = NULL,
            region_subject_out_of = NULL,
            ward_subject_pos_gvt = NULL,
            ward_subject_pos_pvt = NULL,
            council_subject_pos_gvt = NULL,
            council_subject_pos_pvt = NULL,
            region_subject_pos_gvt = NULL,
            region_subject_pos_pvt = NULL,
            submitted_by = NULL,
            submitted_on = NULL
    """
    
    # Add overall_marks if requested
    if clear_overall_marks:
        sql += ", overall_marks = NULL"
    
    # Add WHERE clause
    sql += " WHERE exam_id = %s"
    
    try:
        # Create connection pool
        pool = await aiomysql.create_pool(
            host=DB_CONFIG['host'],
            port=DB_CONFIG['port'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            db=DB_CONFIG['db'],
            minsize=1,
            maxsize=DB_CONFIG['maxsize']
        )
        
        async with pool.acquire() as conn:
            async with conn.cursor() as cursor:
                # Execute the update
                await cursor.execute(sql, (exam_id,))
                affected_rows = cursor.rowcount
                
                # Commit the transaction
                await conn.commit()
                
                return {
                    "status": "success",
                    "message": f"Cleared results for {affected_rows} records in exam {exam_id}",
                    "affected_rows": affected_rows
                }
                
    except Exception as e:
        return {
            "status": "error",
            "message": f"Failed to clear results: {str(e)}"
        }
    finally:
        if 'pool' in locals():
            pool.close()
            await pool.wait_closed()


# if __name__ == "__main__":
#     exam_id="1f0656f7-02b1-6b0c-898c-73f64b350f15"
#     import logging
#     from logging.handlers import RotatingFileHandler

#     # Configure logging to file
#     logger = logging.getLogger(__name__)
#     logger.setLevel(logging.INFO)
#     handler = RotatingFileHandler('exam_results.log', maxBytes=10*1024*1024, backupCount=5)
#     formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
#     handler.setFormatter(formatter)
#     logger.addHandler(handler)
#     logger.info(f"Starting exam results processing for exam_id: {exam_id}")

#     asyncio.run(import_marks_from_excel_OLDER(r"C:\Users\droge\OneDrive\Documents\ENTRY_MBEYA_20250724_223553.xlsm"))

async def main():
    try:
        excel_path = r"C:\Users\droge\OneDrive\Documents\ENTRY_MBEYA_20250724_223553.xlsm"
        exam_id = "1f0656e3-8756-680b-ac24-8d5b3e217521"
        region_name = "Mbeya"
        councils=[
            # "Busokelo",
            # "Chunya",
            # "Mbarali",
            # "Kyela",
            "mbeya cc",
            "Mbeya",
            "Rungwe"
        ]
        for council in councils:
            file_path= await export_to_excel(
                exam_id=exam_id,
                region_name=region_name,
                council_name=council,
                marks_filler="DEFAULT"
            )
            print(f"Generated Excel file: {file_path}")
        # file_path = await export_to_excel(
        #     exam_id=exam_id,
        #     region_name=region_name,
        #     marks_filler="JOANNA K. NHENDE"
        # )
        # print(f"Generated Excel file: {file_path}")
        
        uploaded = await import_marks_from_excel_old(excel_path)
        updates = await calculate_marks_and_ranks(exam_id)
        # print(f"Updated {updates} records for exam_id={exam_id}")
        # # await process_exam_results(exam_id)
    except Exception as e:
        logger.error(f"Main function failed: {e}")
        raise

if __name__ == "__main__":
    import asyncio
    asyncio.run(main())